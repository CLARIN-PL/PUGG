import pandas as pd
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Border, Side


def add_bold_line(writer: pd.ExcelWriter, sheet_name: str, row: int, max_column: int) -> None:
    sheet = writer.sheets[sheet_name]
    border = Border(bottom=Side(border_style="thick"))

    for col in range(1, max_column + 1):
        sheet.cell(row=row, column=col).border = border


def merge_cells(
    writer: pd.ExcelWriter, sheet_name: str, col_idx: int, bold_line: bool = False
) -> None:
    sheet = writer.sheets[sheet_name]
    prev_value = None
    start_row = 2
    max_column = sheet.max_column
    for row in range(2, sheet.max_row + 2):
        cell = sheet.cell(row=row, column=col_idx)
        current_value = cell.value
        if current_value != prev_value:
            if prev_value is not None and row - start_row >= 1:
                sheet.merge_cells(
                    start_row=start_row, start_column=col_idx, end_row=row - 1, end_column=col_idx
                )
                if bold_line:
                    add_bold_line(writer, sheet_name, row - 1, max_column)
            start_row = row
            prev_value = current_value


def format_links(writer: pd.ExcelWriter, sheet_name: str, col_idx: int) -> None:
    sheet = writer.sheets[sheet_name]
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col_idx)
        if cell.value is None:  # last row
            continue
        cell.hyperlink = cell.value
        cell.value = "/".join(cell.value.split("/")[4:])
        cell.style = "Hyperlink"


def color_spans(writer: pd.ExcelWriter, sheet_name: str, col_idx: int, spans: pd.Series) -> None:
    sheet = writer.sheets[sheet_name]
    red = InlineFont(color="00FF0000")
    for i, span in enumerate(spans):
        cell = sheet.cell(row=i + 2, column=col_idx)
        text = cell.value

        rich_text_cell = CellRichText()
        rich_text_cell.append(text[: span[0]])
        rich_text_cell.append(TextBlock(red, text[span[0] : span[1]]))
        rich_text_cell.append(text[span[1] :])
        cell.value = rich_text_cell


def merge_cells_for_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns):
        if col in {"q_p_id", "q_id", "p_id", "question", "text", "gpt_answer", "answer_entities"}:
            if col == "q_p_id":
                merge_cells(writer, sheet_name, col_idx=i + 1, bold_line=True)
            else:
                merge_cells(writer, sheet_name, col_idx=i + 1)


def wrap_text(
    writer: pd.ExcelWriter, sheet_name: str, col_idx: int, len_df: int, height: int = 200
) -> None:
    sheet = writer.sheets[sheet_name]
    row_id_to_resize = None
    counter_standard_cells = 0
    for row in range(2, len_df + 3):
        cell = sheet.cell(row=row, column=col_idx)
        cell.alignment = cell.alignment.copy(wrapText=True)
        if cell.value or row == len_df + 2:
            if row_id_to_resize:
                rd = sheet.row_dimensions[row_id_to_resize]
                rd.height = max(20, height - counter_standard_cells * 20)
            row_id_to_resize = row
            counter_standard_cells = 0
        else:
            rd = sheet.row_dimensions[row]
            rd.height = 20
            counter_standard_cells += 1
